from __future__ import annotations
import logging
from collections import defaultdict

from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from grm.constants import (
    ADMINISTRATIVE_LEVEL_UPLOAD_NO_HEADER_MESSAGE,
    ADMINISTRATIVE_LEVEL_UPLOAD_NOT_FOUND_MESSAGE,
    ADMINISTRATIVE_LEVEL_UPLOAD_ROOT_ALREADY_EXISTS_MESSAGE,
    ADMINISTRATIVE_LEVEL_UPLOAD_ROOT_UNIQUE_MESSAGE,
)
from issues.models import AdministrativeLevel, AdministrativeRegion, Issue

logger = logging.getLogger(__name__)


class AdministrativeRegionProcessor:
    """
    Processor for Excel files containing hierarchical administrative regions.

    Optimized to minimize database queries and correctly count duplicates based on
    hierarchical context rather than just name and level.
    """

    def __init__(self):
        # Cache for existing regions: (name, admin_level_id, parent_id) -> AdministrativeRegion
        self.existing_regions_cache: dict[tuple[str, int, int | None], AdministrativeRegion] = {}

        # Cache for administrative levels
        self.admin_levels_cache: dict[str, AdministrativeLevel] = {}

        # Track regions created in this session
        self.created_regions: dict[tuple[str, int, int | None], AdministrativeRegion] = {}

        # Set to track unique hierarchical paths processed
        self.processed_paths: set[tuple[str, ...]] = set()

        # Statistics
        self.stats = {
            'created': 0,
            'real_duplicates': 0,  # Actual duplicate hierarchical paths
            'existing_reused': 0,  # Existing regions that were reused in hierarchy
            'rows_processed': 0,
            'errors': [],
        }

        # Track the unique root
        self.root_region = None

    def load_caches(self):
        """Pre-load existing data into memory caches."""
        logger.info("Loading administrative levels cache...")
        self.admin_levels_cache = {level.name: level for level in AdministrativeLevel.objects.all()}

        logger.info("Loading existing regions cache...")
        for region in AdministrativeRegion.objects.select_related('administrative_level').all():
            cache_key = (region.name, region.administrative_level.id, region.parent_id)
            self.existing_regions_cache[cache_key] = region

    def get_cache_key(
        self, name: str, admin_level: AdministrativeLevel, parent: AdministrativeRegion | None
    ) -> tuple[str, int, int | None]:
        """Generate cache key for a region."""
        parent_id = parent.id if parent else None
        return (name, admin_level.id, parent_id)

    def find_existing_region(
        self, name: str, admin_level: AdministrativeLevel, parent: AdministrativeRegion | None
    ) -> AdministrativeRegion | None:
        """Find existing region in cache."""
        cache_key = self.get_cache_key(name, admin_level, parent)

        # Check existing regions
        if cache_key in self.existing_regions_cache:
            return self.existing_regions_cache[cache_key]

        # Check newly created regions
        if cache_key in self.created_regions:
            return self.created_regions[cache_key]

        return None

    def create_region(
        self, name: str, admin_level: AdministrativeLevel, parent: AdministrativeRegion | None
    ) -> AdministrativeRegion:
        """Create a new AdministrativeRegion and add it to cache."""
        region = AdministrativeRegion(name=name, administrative_level=admin_level, parent=parent)

        cache_key = self.get_cache_key(name, admin_level, parent)
        self.created_regions[cache_key] = region
        self.stats['created'] += 1

        return region

    def is_duplicate_path(self, row_values: list[str]) -> bool:
        """
        Check if this exact hierarchical path has been processed before.

        This correctly identifies duplicate paths regardless of intermediate processing.
        """
        # Create tuple of non-empty values to represent the complete path
        # Handle None values properly
        path_tuple = tuple(str(value).strip() for value in row_values if value is not None and str(value).strip())

        if path_tuple in self.processed_paths:
            return True

        self.processed_paths.add(path_tuple)
        return False

    def process_excel_row(
        self, row_values: list[str], admin_levels: list[AdministrativeLevel]
    ) -> list[AdministrativeRegion]:
        """
        Process a single Excel row, creating regions as needed.

        Returns:
            List of regions in the hierarchical path
        """
        # Trim and filter empty values while preserving order
        # Handle None values and empty strings properly
        processed_values = []
        processed_levels = []

        for i, value in enumerate(row_values):
            if i >= len(admin_levels):
                break
            # Check for None and empty values properly
            if value is not None and str(value).strip():
                processed_values.append(str(value).strip())
                processed_levels.append(admin_levels[i])
            else:
                # Stop processing when we hit the first empty value
                # This handles partial hierarchies correctly
                break

        if not processed_values:
            return []

        # Check for duplicate path using only the processed (non-empty) values
        if self.is_duplicate_path(processed_values):
            self.stats['real_duplicates'] += 1
            logger.debug(f"Duplicate path skipped: {' -> '.join(processed_values)}")
            return []

        regions = []
        parent_region = None

        # Build hierarchy level by level
        for idx, (name, admin_level) in enumerate(zip(processed_values, processed_levels)):
            if idx == 0:
                # First element = root
                if self.root_region is None:
                    # First time: check if a root already exists in DB
                    existing_root = AdministrativeRegion.objects.filter(
                        administrative_level=admin_level,
                        parent=None,
                    ).first()
                    if existing_root and existing_root.name != name:
                        raise ValidationError(
                            ADMINISTRATIVE_LEVEL_UPLOAD_ROOT_ALREADY_EXISTS_MESSAGE
                            % {"root": existing_root.name, "new": name}
                        )
                    self.root_region = (
                        existing_root
                        or self.find_existing_region(name, admin_level, None)
                        or self.create_region(name, admin_level, None)
                    )
                else:
                    # Already set root, must match
                    if self.root_region.name != name:
                        raise ValidationError(
                            ADMINISTRATIVE_LEVEL_UPLOAD_ROOT_UNIQUE_MESSAGE
                            % {"root": self.root_region.name, "new": name}
                        )
                current_region = self.root_region
            else:
                # Non-root levels as usual
                existing_region = self.find_existing_region(name, admin_level, parent_region)
                if existing_region:
                    current_region = existing_region
                    self.stats['existing_reused'] += 1
                    logger.debug(f"Reusing existing region: {name} (Level: {admin_level.name})")
                else:
                    # Create new region
                    current_region = self.create_region(name, admin_level, parent_region)
                    logger.debug(
                        f"Creating new region: {name} (Level: {admin_level.name}, Parent: {parent_region.name if parent_region else 'None'})"
                    )

            regions.append(current_region)
            parent_region = current_region  # This becomes parent for next level

        return regions

    def bulk_save_regions(self):
        """
        Save all newly created regions to database level by level.

        This ensures parent regions are saved before their children,
        avoiding the 'unsaved related object parent' error.
        """
        if not self.created_regions:
            logger.info("No new regions to save")
            return

        logger.info(f"Saving {len(self.created_regions)} new regions to database...")

        # Group regions by administrative level ID to save in hierarchical order
        regions_by_level = defaultdict(list)

        for region in self.created_regions.values():
            level_id = region.administrative_level.id
            regions_by_level[level_id].append(region)

        # Save regions level by level (from root to leaves)
        total_saved = 0
        for level_id in sorted(regions_by_level.keys()):
            regions_to_save = regions_by_level[level_id]

            logger.info(f"Saving {len(regions_to_save)} regions for level {level_id}")

            # Bulk create regions for this level
            AdministrativeRegion.objects.bulk_create(regions_to_save, batch_size=1000)
            total_saved += len(regions_to_save)

            # Update existing cache with saved regions (they now have IDs)
            for region in regions_to_save:
                cache_key = self.get_cache_key(region.name, region.administrative_level, region.parent)
                self.existing_regions_cache[cache_key] = region

        logger.info(f"Successfully saved {total_saved} regions across {len(regions_by_level)} levels")

    def clean_unused_regions(self):
        """Delete AdministrativeRegion objects that are not being used."""
        from django.db.models import Exists, OuterRef

        # Find regions that are not referenced by any Issue
        related_qs = AdministrativeRegion.objects.annotate(
            has_issue=Exists(Issue.objects.filter(administrative_region=OuterRef("pk")))
        )

        to_delete = related_qs.filter(has_issue=False)
        not_deleted_count = related_qs.filter(has_issue=True).count()

        deleted_count, _ = to_delete.delete()

        logger.info(f"Deleted {deleted_count} unused regions, {not_deleted_count} regions kept (in use)")

        return deleted_count, not_deleted_count

    def process_excel(self, file_obj) -> tuple[int, int, int]:
        """
        Main method to process Excel file and create AdministrativeRegions.

        Returns:
            Tuple of (created_count, duplicate_count)
        """
        logger.info("Starting Excel processing...")

        # Load caches
        self.load_caches()

        try:
            # Load workbook
            wb = load_workbook(file_obj)
            ws = wb.active

            # Check header row exists
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
            if not headers or all(h is None for h in headers):
                raise ValidationError(ADMINISTRATIVE_LEVEL_UPLOAD_NO_HEADER_MESSAGE)

            # Get header row to determine administrative levels
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            level_names = [str(cell).strip() for cell in header_row if cell]

            # Validate and get administrative levels
            admin_levels = []
            for level_name in level_names:
                if level_name not in self.admin_levels_cache:
                    raise ValidationError(ADMINISTRATIVE_LEVEL_UPLOAD_NOT_FOUND_MESSAGE % {'level': level_name})
                admin_levels.append(self.admin_levels_cache[level_name])

            logger.info(f"Processing hierarchy: {' -> '.join(level_names)}")

            # Clean unused regions before processing
            deleted_count, not_deleted_count = self.clean_unused_regions()

            # Process data rows
            with transaction.atomic():
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    # Skip completely empty rows
                    if not any(row):
                        continue

                    try:
                        # Convert row to list and ensure we have enough values
                        row_values = list(row)
                        if len(row_values) < len(admin_levels):
                            row_values.extend([None] * (len(admin_levels) - len(row_values)))

                        # Process this row
                        self.process_excel_row(row_values[: len(admin_levels)], admin_levels)
                        self.stats['rows_processed'] += 1

                        # Log progress every 1000 rows
                        if self.stats['rows_processed'] % 1000 == 0:
                            logger.info(f"Processed {self.stats['rows_processed']} rows...")

                    except Exception as e:
                        error_msg = f"Error processing row {row_idx}: {list(row)} - {str(e)}"
                        logger.error(error_msg)
                        self.stats['errors'].append(error_msg)
                        continue

                # Save all new regions
                self.bulk_save_regions()

            logger.info("Excel processing completed successfully")
            logger.info(
                f"Statistics: Created: {self.stats['created']}, "
                f"Real duplicates: {self.stats['real_duplicates']}, "
                f"Existing reused: {self.stats['existing_reused']}, "
                f"Errors: {len(self.stats['errors'])}"
            )

            return self.stats['created'], self.stats['real_duplicates'], not_deleted_count

        except Exception as e:
            logger.error(f"Excel processing failed: {str(e)}")
            raise ValidationError(str(e))
