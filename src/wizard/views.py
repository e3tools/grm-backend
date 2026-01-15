from __future__ import annotations

import logging

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.postgres.aggregates import ArrayAgg
from django.db.models import Exists, OuterRef, Q
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
from django.views.generic import ListView, TemplateView, UpdateView, View
from django.views.generic.edit import FormView
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from dashboard.mixins import LoginRequiredAndAJAXRequestMixin
from dashboard.models import Project
from grm.constants import MAP_CONFIDENTIALITY_LEVEL, MAP_REDIRECTION_PROTOCOL
from issues.models import (
    AdministrativeLevel,
    AdministrativeRegion,
    Citizen,
    CitizenAgeGroup,
    CitizenGroup,
    Component,
    Issue,
    IssueCategory,
    IssueDepartment,
    IssueDepartmentAdministrativeLevel,
    IssueStatus,
    IssueType,
    SubComponent,
)
from wizard import constants as cons
from wizard.forms import (
    DEFAULT_CITIZEN_AGE_GROUPS,
    ISSUE_STATUS_DEFINITIONS,
    AdministrativeLevelFormSet,
    ExistingCitizenAgeGroupFormSet,
    ExistingCitizenGroupFormSet,
    ExistingComponentFormSet,
    ExistingIssueStatusFormSet,
    ExistingIssueTypeFormSet,
    IssueCategoryFormSet,
    IssueDepartmentFormSet,
    NewCitizenAgeGroupFormSet,
    NewCitizenGroupFormSet,
    NewComponentFormSet,
    NewIssueStatusFormSet,
    NewIssueTypeFormSet,
    ProjectForm,
    SubComponentFormSet,
    UploadAdministrativeRegionForm,
)
from wizard.models import WizardSection
from wizard.registry import (
    get_next_step,
    get_step_by_name,
    get_total_steps,
    register_wizard_step,
)
from wizard.utils import AdministrativeRegionProcessor

logger = logging.getLogger(__name__)


class CustomizationWizardView(LoginRequiredMixin, TemplateView):
    template_name = "wizard/grm_customization.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sections = list(WizardSection.objects.values_list('id', flat=True))
        total_steps = len(sections)
        in_progress_section = WizardSection.objects.filter(status=cons.IN_PROGRESS_CHOICE).first()
        ips_id = in_progress_section.id if in_progress_section else None
        in_progress_step = sections.index(ips_id) + 1 if ips_id else total_steps

        current_step = self.request.GET.get('step')
        if current_step and current_step.isdigit():
            current_step = min(int(current_step), in_progress_step)
        else:
            current_step = in_progress_step

        context['current_step'] = current_step
        return context


class WizardSectionListView(LoginRequiredAndAJAXRequestMixin, ListView):
    template_name = "wizard/wizard_sections.html"
    context_object_name = "wizard_sections"
    model = WizardSection

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['step'] = int(self.request.GET.get("step"))
        return context


class WizardFormView(LoginRequiredAndAJAXRequestMixin, FormView):
    form_class = ProjectForm
    template_name = "wizard/form.html"
    step_name = None  # Define in each subclass
    step = None

    def dispatch(self, request, *args, **kwargs):
        self.step = self.get_step_config()['step']
        return super().dispatch(request, *args, **kwargs)

    def get_step_config(self):
        """Get step configuration from registry."""
        if not self.step_name:
            raise ValueError("step_name must be defined in the view class")
        return get_step_by_name(self.step_name)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['step'] = self.step
        context['card_title'] = _("Configuration Options")
        context['toast_title'] = cons.NOT_PERMITTED_TEXT
        context['toast_message'] = cons.ITEM_TOAST_ERROR_MESSAGE
        return context

    def get_success_url(self):
        """Get the URL for the next step."""
        next_step = get_next_step(self.step_name)
        if next_step:
            return reverse(f"wizard:setup_step_{next_step['step']}")
        return reverse("wizard:customization_wizard")

    def get_current_section(self):
        """Get the current WizardSection based on step number."""
        return WizardSection.objects.get(step=self.step)

    def get_next_section(self):
        """Get the next WizardSection."""
        next_step = get_next_step(self.step_name)
        if next_step:
            return WizardSection.objects.filter(step=next_step['step']).first()
        return None

    def update_status(self, only_current_step=False, status=cons.COMPLETED_CHOICE):
        """Update wizard section status."""
        current_section = self.get_current_section()
        current_section.status = status
        current_section.save(update_fields=['status'])

        if not only_current_step:
            next_section = self.get_next_section()
            if next_section and next_section.status == cons.NOT_STARTED_CHOICE:
                next_section.status = cons.IN_PROGRESS_CHOICE
                next_section.save(update_fields=['status'])

    def form_valid(self, form):
        form.save()
        self.update_status()
        return super().form_valid(form)


@register_wizard_step(cons.PROJECT_CHOICE)
class ProjectUpdateView(WizardFormView, UpdateView):
    step_name = cons.PROJECT_CHOICE

    def get_object(self, queryset=None):
        obj = Project.objects.first()
        if not obj:
            obj = Project()
        return obj


@register_wizard_step(cons.ADMINISTRATIVE_LEVELS_CHOICE)
class AdministrativeLevelsFormView(WizardFormView):
    form_class = AdministrativeLevelFormSet
    template_name = "wizard/formset.html"
    step_name = cons.ADMINISTRATIVE_LEVELS_CHOICE

    def get_form(self, form_class=None):
        queryset = AdministrativeLevel.objects.annotate(
            restricted_deletion=(
                Exists(Issue.objects.filter(administrative_region__administrative_level=OuterRef("pk")))
                | Exists(IssueDepartmentAdministrativeLevel.objects.filter(administrative_level=OuterRef("pk")))
            )
        )
        return self.form_class(queryset=queryset, **self.get_form_kwargs())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['formset'] = context['form']  # Aliases for clarity in the template
        context['formset_label'] = _('Administrative Level Names')
        return context


class DownloadRegionsSampleView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = str(cons.ADMINISTRATIVE_LEVEL_EXCEL_WORKBOOK_TITLE)

        # 1. Get levels
        levels = AdministrativeLevel.objects.all()
        headers = [level.name for level in levels]

        # 2. Writing headers
        ws.append(headers)

        # 3. Building rows with existing regions
        rows = self._build_region_rows(levels)
        for row in rows:
            ws.append(row)

        # 4. Adjust column width
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="administrative_levels_sample.xlsx"'
        wb.save(response)
        return response

    def _build_region_rows(self, levels):
        """
        Build the rows for Excel:
        Each row represents a chain of AdministrativeRegions from level 1 to the last.
        """
        if not levels.exists():
            return []

        rows = []

        # get root (there can only be one per model)
        root_region = AdministrativeRegion.objects.filter(parent__isnull=True).first()

        if root_region:
            self._add_region_recursive(root_region, [root_region.name], rows, levels, 1)

        return rows

    def _add_region_recursive(self, region, current_path, rows, levels, level_index):
        """
        Recursively traverses children and builds rows with full paths
        """
        if level_index == len(levels):  # last level reached
            rows.append(current_path + [""] * (len(levels) - len(current_path)))
            return

        children = list(region.children.all())
        if children:
            for child in children:
                self._add_region_recursive(child, current_path + [child.name], rows, levels, level_index + 1)
        else:
            # fill up to the number of levels with empty cells
            rows.append(current_path + [""] * (len(levels) - len(current_path)))


@register_wizard_step(cons.ADMINISTRATIVE_REGIONS_CHOICE)
class AdministrativeRegionFormView(WizardFormView):
    template_name = "wizard/regions.html"
    form_class = UploadAdministrativeRegionForm
    step_name = cons.ADMINISTRATIVE_REGIONS_CHOICE

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["regions_summary"] = AdministrativeLevel.get_regions_summary()
        current_section = WizardSection.objects.all()[self.step - 1]
        context["section_status"] = current_section.status
        return context

    def form_valid(self, form):
        file = form.cleaned_data["file"]

        try:
            """Parse Excel and create AdministrativeRegions in a hierarchical structure."""

            processor = AdministrativeRegionProcessor()

            created_count, duplicate_count, not_deleted_count = processor.process_excel(file)

            # Handle any processing errors
            if processor.stats['errors']:
                error_summary = f"Processing completed with {len(processor.stats['errors'])} errors. "
                error_summary += "Check logs for details."
                logger.warning(error_summary)

                for error in processor.stats['errors']:
                    messages.error(
                        self.request,
                        error,
                        extra_tags="danger",
                    )

            if created_count:
                messages.success(
                    self.request,
                    cons.ADMINISTRATIVE_LEVEL_UPLOAD_SUCCESS_MESSAGE % {"count": created_count},
                    extra_tags="success",
                )
                self.update_status(only_current_step=True)
            if duplicate_count:
                messages.warning(
                    self.request,
                    cons.ADMINISTRATIVE_LEVEL_UPLOAD_DUPLICATES_MESSAGE % {"count": duplicate_count},
                    extra_tags="warning",
                )
            if not_deleted_count > 0:
                messages.warning(
                    self.request,
                    cons.ADMINISTRATIVE_LEVEL_UPLOAD_UNCHANGEABLE_MESSAGE % {"count": not_deleted_count},
                    extra_tags="warning",
                )
        except Exception as e:
            messages.error(
                self.request,
                _("There was an error processing the Excel file: %(error)s") % {"error": str(e)},
                extra_tags="danger",
            )

        if not AdministrativeRegion.objects.exists():
            messages.warning(
                self.request,
                cons.ADMINISTRATIVE_LEVEL_UPLOAD_DELETE_MESSAGE,
                extra_tags="warning",
            )
            self.update_status(only_current_step=True, status=cons.IN_PROGRESS_CHOICE)

        context = {"msg": render(self.request, "common/messages.html").content.decode("utf-8")}
        return JsonResponse(context, safe=False)


class NextStepView(LoginRequiredAndAJAXRequestMixin, View):
    def post(self, request, *args, **kwargs):
        step = self.kwargs["step"]
        sections = WizardSection.objects.all()
        current_section = sections[step - 1]
        if current_section.status == cons.COMPLETED_CHOICE:
            WizardSection.objects.filter(id=current_section.id + 1, status=cons.NOT_STARTED_CHOICE).update(
                status=cons.IN_PROGRESS_CHOICE
            )
            step += 1
        return JsonResponse({"step": step}, safe=False)


@register_wizard_step(cons.DEPARTMENTS_CHOICE)
class IssueDepartmentsFormView(WizardFormView):
    form_class = IssueDepartmentFormSet
    template_name = "wizard/formset.html"
    step_name = cons.DEPARTMENTS_CHOICE

    def get_form(self, form_class=None):
        queryset = IssueDepartment.objects.annotate(
            restricted_deletion=Exists(
                IssueCategory.objects.filter(
                    Q(assigned_department__department=OuterRef("pk"))
                    | Q(assigned_appeal_department__department=OuterRef("pk"))
                    | Q(assigned_escalation_department__department=OuterRef("pk"))
                )
            )
        )
        return self.form_class(queryset=queryset, **self.get_form_kwargs())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['formset'] = context['form']
        context['formset_label'] = _('Departments')
        context['two_fields_by_row'] = True
        return context


@register_wizard_step(cons.ISSUE_TYPES_CHOICE)
class IssueTypesFormView(WizardFormView):
    form_class = NewIssueTypeFormSet
    template_name = "wizard/formset.html"
    step_name = cons.ISSUE_TYPES_CHOICE

    def get_form(self, form_class=None):
        queryset = IssueType.objects.annotate(
            restricted_deletion=Exists(IssueCategory.objects.filter(parent__parent=OuterRef("pk")))
        )

        if queryset.exists():
            return ExistingIssueTypeFormSet(queryset=queryset, **self.get_form_kwargs())
        else:
            kwargs = self.get_form_kwargs()
            kwargs['queryset'] = IssueType.objects.none()
            return self.form_class(**kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['formset'] = context['form']
        context['formset_label'] = _('Issue Types')
        return context


@register_wizard_step(cons.CATEGORIES_CHOICE)
class IssueCategoriesFormView(WizardFormView):
    form_class = IssueCategoryFormSet
    template_name = "wizard/formset.html"
    step_name = cons.CATEGORIES_CHOICE

    def get_form(self, form_class=None):
        queryset = IssueCategory.objects.annotate(
            restricted_deletion=Exists(Issue.objects.filter(category=OuterRef("pk")))
        ).select_related('parent')
        return self.form_class(queryset=queryset, **self.get_form_kwargs())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['formset'] = context['form']
        context['formset_label'] = _('Categories')
        context['two_fields_by_row'] = True
        return context


@register_wizard_step(cons.ISSUE_STATUS_CHOICE)
class ResolutionProcessFormView(WizardFormView):
    form_class = NewIssueStatusFormSet
    template_name = "wizard/static_formset.html"
    step_name = cons.ISSUE_STATUS_CHOICE

    def get_form(self, form_class=None):
        queryset = IssueStatus.objects.all()

        if queryset.exists():
            return ExistingIssueStatusFormSet(queryset=queryset, **self.get_form_kwargs())
        else:
            initial_data = [{'name': item['name']} for item in ISSUE_STATUS_DEFINITIONS.values()]
            kwargs = self.get_form_kwargs()
            kwargs['initial'] = initial_data
            kwargs['queryset'] = IssueStatus.objects.none()

            return self.form_class(**kwargs)

    def form_valid(self, formset):
        instances = formset.save(commit=False)

        for instance, flag in zip(instances, ISSUE_STATUS_DEFINITIONS.keys()):
            if not instance.pk:
                instance.initial_status = flag == "initial_status"
                instance.open_status = flag == "open_status"
                instance.final_status = flag == "final_status"
                instance.rejected_status = flag == "rejected_status"
            instance.save()

        self.update_status()
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['formset'] = context['form']  # Aliases for clarity in the template
        context['formset_label'] = _('Issue Status')
        return context


@register_wizard_step(cons.CITIZEN_AGE_GROUPS_CHOICE)
class CitizenAgeGroupsFormView(WizardFormView):
    form_class = NewCitizenAgeGroupFormSet
    template_name = "wizard/formset.html"
    step_name = cons.CITIZEN_AGE_GROUPS_CHOICE

    def get_form(self, form_class=None):
        queryset = CitizenAgeGroup.objects.annotate(
            restricted_deletion=Exists(Citizen.objects.filter(age_group=OuterRef("pk")))
        )

        if queryset.exists():
            return ExistingCitizenAgeGroupFormSet(queryset=queryset, **self.get_form_kwargs())
        else:
            initial_data = [{'name': name} for name in DEFAULT_CITIZEN_AGE_GROUPS]
            kwargs = self.get_form_kwargs()
            kwargs['initial'] = initial_data
            kwargs['queryset'] = IssueStatus.objects.none()

            return self.form_class(**kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['formset'] = context['form']  # Aliases for clarity in the template
        context['formset_label'] = _('Citizen Age Groups')
        return context


@register_wizard_step(cons.CITIZEN_GROUPS_CHOICE)
class CitizenGroupsFormView(WizardFormView):
    form_class = NewCitizenGroupFormSet
    template_name = "wizard/formset.html"
    step_name = cons.CITIZEN_GROUPS_CHOICE
    queryset = None

    def get_form(self, form_class=None):
        self.queryset = CitizenGroup.objects.annotate(
            restricted_deletion=Exists(Citizen.objects.filter(age_group=OuterRef("pk")))
        )

        if self.queryset.exists():
            self.form_class = ExistingCitizenGroupFormSet

        return self.form_class(queryset=self.queryset, **self.get_form_kwargs())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['formset'] = context['form']  # Aliases for clarity in the template
        context['formset_label'] = _('Citizen Groups')
        context['skip'] = False if self.queryset else True
        return context


@register_wizard_step(cons.COMPONENTS_CHOICE)
class ComponentAndSubComponentFormView(WizardFormView):
    form_class = NewComponentFormSet
    template_name = "wizard/nested_formset.html"
    step_name = cons.COMPONENTS_CHOICE

    def get_form(self, form_class=None):
        """Get the appropriate formset with subformsets and restricted_deletion annotations."""

        # Annotate Components with restricted_deletion
        queryset = Component.objects.annotate(
            # Check if Component is directly referenced by Issues
            restricted_deletion=Exists(Issue.objects.filter(component=OuterRef("pk")))
        )

        if queryset.exists():
            self.form_class = ExistingComponentFormSet

        formset = self.form_class(queryset=queryset, **self.get_form_kwargs())

        used_subcomponents = SubComponent.objects.filter(issues__isnull=False).distinct().values_list('id', flat=True)

        # Initialize subformsets with POST data if available
        if self.request.method == 'POST':
            # Annotate SubComponents with restricted_deletion
            subcomponents_queryset = SubComponent.objects.annotate(
                restricted_deletion=Exists(Issue.objects.filter(sub_component=OuterRef("pk")))
            )
            formset.subformsets = []
            for i, form in enumerate(formset.forms):
                if form.instance.pk:
                    qs = subcomponents_queryset.filter(parent=form.instance)
                else:
                    qs = subcomponents_queryset.none()

                subformset = SubComponentFormSet(
                    self.request.POST,
                    instance=form.instance if form.instance.pk else None,
                    queryset=qs,
                    prefix=f'subcomponent_form-{i}',
                )
                # Link parent form for validation context
                for subform in subformset.forms:
                    subform.parent_form = form
                    subform.restricted_deletion = subform.instance.id in used_subcomponents
                formset.subformsets.append(subformset)
        else:
            # Link parent forms for GET requests too
            for form, subformset in zip(formset.forms, formset.subformsets):
                for subform in subformset.forms:
                    subform.parent_form = form
                    subform.restricted_deletion = subform.instance.id in used_subcomponents

        return formset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['formset'] = context['form']
        context["formset_label"] = _("Components and Subcomponents")
        context["toast_title"] = cons.NOT_PERMITTED_TEXT
        context["toast_message"] = cons.ITEM_TOAST_ERROR_MESSAGE
        return context

    def form_valid(self, formset):
        """Save components and their subcomponents."""
        instances = formset.save(commit=False)

        # Save each component and its subcomponents
        for instance in instances:
            instance.save()

        # Handle deleted components
        for obj in formset.deleted_objects:
            obj.delete()

        return super().form_valid(formset)


@register_wizard_step(cons.SUMMARY_CHOICE)
class SummaryView(LoginRequiredAndAJAXRequestMixin, TemplateView):
    template_name = "wizard/summary.html"
    step_name = cons.SUMMARY_CHOICE

    def get_step_number(self):
        return get_step_by_name(self.step_name)['step']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['step'] = self.get_step_number()
        context['card_title'] = cons.SUMMARY_DISPLAY
        context['disabled_submit'] = self._is_submit_disabled()
        context['summary'] = self._build_summary()
        return context

    def _is_submit_disabled(self):
        """Check if all previous steps are completed."""
        current_step = self.get_step_number()
        return WizardSection.objects.exclude(step=current_step).exclude(status=cons.COMPLETED_CHOICE).exists()

    def _build_summary(self):
        """Build complete summary data for all steps dynamically."""
        from wizard.registry import get_all_wizard_steps

        summary = []
        wizard_steps = get_all_wizard_steps()

        for step_config in sorted(wizard_steps.values(), key=lambda x: x['step']):
            if step_config.get('name') == cons.SUMMARY_CHOICE:
                continue

            section_name = None
            for name, config in wizard_steps.items():
                if config['step'] == step_config['step']:
                    section_name = name
                    break

            if section_name:
                method_name = f"_get_{section_name}_summary"
                if hasattr(self, method_name):
                    summary.append(getattr(self, method_name)())

        return summary

    @staticmethod
    def _get_project_summary():
        """Get project information summary."""
        project = Project.objects.first()
        return {
            "title": cons.PROJECT_DISPLAY,
            "data": [
                {
                    "fields": [
                        {"label": _("Name"), "value": project.name if project else ""},
                        {"label": _("Description"), "value": project.description if project else ""},
                    ]
                }
            ],
        }

    @staticmethod
    def _get_administrative_levels_summary():
        """Get administrative levels summary."""
        data = [{"fields": [{"label": _("Name"), "value": level.name}]} for level in AdministrativeLevel.objects.all()]
        return {"title": cons.ADMINISTRATIVE_LEVELS_DISPLAY, "data": data}

    @staticmethod
    def _get_administrative_regions_summary():
        """Get administrative regions summary."""
        return {
            "title": cons.ADMINISTRATIVE_REGIONS_DISPLAY,
            "regions_summary": AdministrativeLevel.get_regions_summary(),
        }

    @staticmethod
    def _get_departments_summary():
        """Get departments summary."""
        departments = (
            IssueDepartmentAdministrativeLevel.objects.values('department__name')
            .annotate(administrative_levels=ArrayAgg('administrative_level__name', distinct=True))
            .order_by('department__name')
        )

        data = [
            {
                "fields": [
                    {"label": _("Name"), "value": dept["department__name"]},
                    {"label": _("Administrative levels"), "value": ', '.join(dept["administrative_levels"])},
                ]
            }
            for dept in departments
        ]

        return {"title": cons.DEPARTMENTS_DISPLAY, "data": data}

    @staticmethod
    def _get_issue_types_summary():
        """Get issue types summary."""

        data = [
            {
                "fields": [
                    {"label": _("Name"), "value": issue_type.name},
                    {"label": _("Subtypes"), "value": ', '.join(issue_type.children.values_list("name", flat=True))},
                ]
            }
            for issue_type in IssueType.objects.prefetch_related("children")
        ]

        return {"title": cons.ISSUE_TYPES_DISPLAY, "data": data}

    @staticmethod
    def _get_categories_summary():
        """Get issue categories summary."""
        categories = IssueCategory.objects.select_related(
            'parent',
            'assigned_department__department',
            'assigned_appeal_department__department',
            'assigned_escalation_department__department',
        )

        data = [
            {
                "fields": [
                    {"label": _("Name"), "value": cat.name},
                    {"label": _("Abbreviation"), "value": cat.abbreviation},
                    {"label": _("Subtype"), "value": cat.parent.name if cat.parent else ""},
                    {"label": _("Department"), "value": cat.assigned_department.department.name},
                    {"label": _("Appeal department"), "value": cat.assigned_appeal_department.department.name},
                    {"label": _("Escalation department"), "value": cat.assigned_escalation_department.department.name},
                    {
                        "label": _("Confidentiality level"),
                        "value": MAP_CONFIDENTIALITY_LEVEL.get(cat.confidentiality_level, ""),
                    },
                    {
                        "label": _("Redirection protocol"),
                        "value": MAP_REDIRECTION_PROTOCOL.get(cat.redirection_protocol, ""),
                    },
                ]
            }
            for cat in categories
        ]

        return {"title": cons.CATEGORIES_DISPLAY, "data": data}

    def _get_issue_status_summary(self):
        """Get issue status summary."""
        status_fields = []
        for status in IssueStatus.objects.all():
            label = self._get_status_label(status)
            status_fields.append({"label": label, "value": status.name})

        return {"title": cons.ISSUE_STATUS_DISPLAY, "data": [{"fields": status_fields}]}

    @staticmethod
    def _get_status_label(status):
        """Determine the label for a status based on its flags."""
        for flag in ('initial_status', 'open_status', 'rejected_status', 'final_status'):
            if getattr(status, flag, False):
                return ISSUE_STATUS_DEFINITIONS.get(flag, {}).get('label', '')
        return ''

    @staticmethod
    def _get_citizen_age_groups_summary():
        """Get citizen age groups summary."""
        data = [{"fields": [{"label": _("Name"), "value": group.name}]} for group in CitizenAgeGroup.objects.all()]

        return {"title": cons.CITIZEN_AGE_GROUPS_DISPLAY, "data": data}

    @staticmethod
    def _get_citizen_groups_summary():
        """Get citizen groups summary."""
        data = [
            {
                "fields": [
                    {"label": _("Name"), "value": group.name},
                    {"label": _("Type"), "value": cons.MAP_CITIZEN_GROUP.get(group.type, "")},
                ]
            }
            for group in CitizenGroup.objects.all()
        ]

        return {"title": cons.CITIZEN_GROUPS_DISPLAY, "data": data}

    @staticmethod
    def _get_components_summary():
        """Get components and subcomponents summary."""
        components = Component.objects.prefetch_related('subcomponent_set')

        data = [
            {
                "fields": [
                    {"label": _("Name"), "value": component.name},
                    {"label": _("Description"), "value": component.description},
                ],
                "subcomponents": [
                    {
                        "fields": [
                            {"label": _("Name"), "value": sub.name},
                            {"label": _("Description"), "value": sub.description},
                        ]
                    }
                    for sub in component.subcomponent_set.all()
                ],
            }
            for component in components
        ]

        return {"title": cons.COMPONENTS_DISPLAY, "data": data}

    def post(self, request, *args, **kwargs):
        """Handle summary completion and wizard finalization."""
        if self._can_complete_setup():
            self._mark_setup_complete()
            redirect_url = reverse("dashboard:diagnostics:home")
            message_context = {}
        else:
            redirect_url = None
            messages.error(
                request,
                _("The setup cannot be completed until all previous steps are completed."),
                extra_tags="danger",
            )
            message_context = {"msg": render(request, "common/messages.html").content.decode("utf-8")}

        context = {
            **message_context,
            "redirect_url": redirect_url,
        }
        return JsonResponse(context, safe=False)

    def _can_complete_setup(self):
        """Check if all required steps are completed."""
        current_step = self.get_step_number()
        completed_count = WizardSection.objects.exclude(step=current_step).filter(status=cons.COMPLETED_CHOICE).count()
        return completed_count == get_total_steps() - 1

    def _mark_setup_complete(self):
        """Mark the current wizard section as completed."""
        current_section = WizardSection.objects.get(step=self.get_step_number())
        current_section.status = cons.COMPLETED_CHOICE
        current_section.save(update_fields=['status'])
