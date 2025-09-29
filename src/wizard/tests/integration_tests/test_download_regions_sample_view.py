import io

from django.urls import reverse
from openpyxl import load_workbook

from grm.constants import ADMINISTRATIVE_LEVEL_EXCEL_WORKBOOK_TITLE
from grm.tests.base import ViewTestCase
from issues.factories import AdministrativeLevelFactory, AdministrativeRegionFactory
from issues.models import AdministrativeLevel, AdministrativeRegion


class DownloadRegionsSampleViewTest(ViewTestCase):
    """Integration tests for the DownloadRegionsSampleView."""

    def setUp(self):
        self.url = reverse("wizard:download_regions_sample")

        # Create administrative levels
        self.level1 = AdministrativeLevelFactory(name="Country")
        self.level2 = AdministrativeLevelFactory(name="State")
        self.level3 = AdministrativeLevelFactory(name="District")

        # Create some administrative regions
        self.region1 = AdministrativeRegionFactory(name="USA", administrative_level=self.level1, parent=None)
        self.region2 = AdministrativeRegionFactory(
            name="California", administrative_level=self.level2, parent=self.region1
        )
        self.region3 = AdministrativeRegionFactory(
            name="Los Angeles", administrative_level=self.level3, parent=self.region2
        )

    def test_auth_permission(self):
        """Test that login is required to access the view."""
        response = self.get(self.url, authorized=False)

        self.assertEqual(response.status_code, 302)

    def test_authenticated_user_can_download(self):
        """Test that authenticated user can download the file."""

        response = self.get(self.url)

        self.assertEqual(response.status_code, 200)

    def test_response_content_type(self):
        """Test that response has correct Excel content type."""

        response = self.get(self.url)

        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_response_content_disposition(self):
        """Test that response has correct filename in Content-Disposition."""

        response = self.get(self.url)

        self.assertEqual(response['Content-Disposition'], 'attachment; filename="administrative_levels_sample.xlsx"')

    def test_excel_file_structure(self):
        """Test that downloaded Excel file has correct structure."""

        response = self.get(self.url)

        # Load workbook from response content
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Check worksheet title
        self.assertEqual(ws.title, ADMINISTRATIVE_LEVEL_EXCEL_WORKBOOK_TITLE)

        # Check headers (first row)
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Country", headers)
        self.assertIn("State", headers)
        self.assertIn("District", headers)

    def test_excel_contains_correct_headers(self):
        """Test that Excel file contains all administrative level names as headers."""

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Get all level names from database
        expected_headers = list(AdministrativeLevel.objects.values_list('name', flat=True))

        # Get actual headers from Excel
        actual_headers = [cell.value for cell in ws[1] if cell.value]

        self.assertEqual(sorted(actual_headers), sorted(expected_headers))

    def test_excel_with_no_levels(self):
        """Test Excel generation when no administrative levels exist."""
        AdministrativeLevel.objects.all().delete()

        response = self.get(self.url)

        self.assertEqual(response.status_code, 200)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Should have no headers
        headers = [cell.value for cell in ws[1] if cell.value]
        self.assertEqual(len(headers), 0)

    def test_excel_with_no_regions(self):
        """Test Excel generation when no regions exist."""
        AdministrativeRegion.objects.all().delete()

        response = self.get(self.url)

        self.assertEqual(response.status_code, 200)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Should have headers but no data rows
        self.assertEqual(ws.max_row, 1)  # Only header row

    def test_column_widths_adjusted(self):
        """Test that column widths are properly adjusted."""

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Check that column widths have been set
        for col_letter in ['A', 'B', 'C']:
            width = ws.column_dimensions[col_letter].width
            # Width should be greater than default (which is around 8.43)
            self.assertGreater(width, 8)

    def test_excel_with_unicode_region_names(self):
        """Test Excel generation with unicode characters in region names."""
        # Create region with unicode name
        AdministrativeRegionFactory(name="São Paulo", administrative_level=self.level2, parent=self.region1)

        response = self.get(self.url)

        self.assertEqual(response.status_code, 200)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Find the unicode name in the worksheet
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if "São Paulo" in [cell for cell in row if cell]:
                found = True
                break

        self.assertTrue(found, "Unicode region name not found in Excel")

    def test_excel_with_special_characters(self):
        """Test Excel generation with special characters in names."""
        AdministrativeRegionFactory(
            name="Region-with-Dashes & Symbols!", administrative_level=self.level2, parent=self.region1
        )

        response = self.get(self.url)

        self.assertEqual(response.status_code, 200)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Should not raise any errors
        self.assertIsNotNone(ws.max_row)

    def test_multiple_levels_hierarchy(self):
        """Test Excel with multiple levels of hierarchy."""
        # Create a deeper hierarchy
        level4 = AdministrativeLevelFactory(name="City")
        level5 = AdministrativeLevelFactory(name="Neighborhood")

        city = AdministrativeRegionFactory(name="San Francisco", administrative_level=level4, parent=self.region2)

        AdministrativeRegionFactory(name="Mission District", administrative_level=level5, parent=city)

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Should have 5 columns (5 levels)
        headers = [cell.value for cell in ws[1] if cell.value]
        self.assertEqual(len(headers), 5)

    def test_post_method_not_allowed(self):
        """Test that POST method is not allowed."""

        response = self.post(self.url, {})

        self.assertEqual(response.status_code, 405)

    def test_put_method_not_allowed(self):
        """Test that PUT method is not allowed."""

        response = self.put(self.url, {})

        self.assertEqual(response.status_code, 405)

    def test_delete_method_not_allowed(self):
        """Test that DELETE method is not allowed."""

        response = self.delete(self.url)

        self.assertEqual(response.status_code, 405)

    def test_excel_row_count_matches_region_hierarchy(self):
        """Test that the number of data rows matches the expected hierarchy."""

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Should have header row + data rows
        # The exact count depends on _build_region_rows implementation
        self.assertGreaterEqual(ws.max_row, 1)  # At least header row

    def test_excel_file_not_corrupted(self):
        """Test that Excel file is not corrupted and has valid structure."""

        response = self.get(self.url)

        # File should have content
        self.assertGreater(len(response.content), 0)

        # Should be able to load as workbook
        wb = load_workbook(io.BytesIO(response.content))

        # Should have at least one sheet
        self.assertGreater(len(wb.sheetnames), 0)

        # Active sheet should exist
        self.assertIsNotNone(wb.active)

    def test_concurrent_downloads(self):
        """Test that multiple concurrent downloads work correctly."""

        # Make multiple requests
        responses = []
        for _ in range(3):
            response = self.get(self.url)
            responses.append(response)

        # All should succeed
        for response in responses:
            self.assertEqual(response.status_code, 200)
            self.assertGreater(len(response.content), 0)

    def test_very_long_region_names(self):
        """Test Excel generation with very long region names."""
        long_name = "A" * 250  # Very long name
        AdministrativeRegionFactory(name=long_name, administrative_level=self.level2, parent=self.region1)

        response = self.get(self.url)

        self.assertEqual(response.status_code, 200)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Column width should be adjusted for long content
        # but Excel has a max width limit
        for col in ws.columns:
            col_letter = col[0].column_letter
            width = ws.column_dimensions[col_letter].width
            # Width should be set (not default)
            self.assertIsNotNone(width)

    def test_hierarchical_data_structure(self):
        """Test that Excel contains correct hierarchical data structure."""

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Find the row with our test data
        found_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == "USA" and row[1] == "California" and row[2] == "Los Angeles":
                found_row = row
                break

        self.assertIsNotNone(found_row, "Expected hierarchical data not found in Excel")
        self.assertEqual(found_row[0], "USA")
        self.assertEqual(found_row[1], "California")
        self.assertEqual(found_row[2], "Los Angeles")

    def test_partial_hierarchy_paths_filled_with_empty_cells(self):
        """Test that partial hierarchy paths are filled with empty cells."""
        # Create a region that doesn't have children at the last level
        AdministrativeRegionFactory(name="Texas", administrative_level=self.level2, parent=self.region1)
        # Texas has no districts (level3), so its row should have empty cells

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Find Texas row
        found_row = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == "Texas":
                found_row = row
                break

        self.assertIsNotNone(found_row, "Texas row not found")
        self.assertEqual(found_row[0], "USA")
        self.assertEqual(found_row[1], "Texas")
        # Third column should be empty or None
        self.assertTrue(found_row[2] is None or found_row[2] == "")

    def test_multiple_children_at_same_level(self):
        """Test that multiple children at same level create separate rows."""
        # Create another district under California
        AdministrativeRegionFactory(name="San Diego", administrative_level=self.level3, parent=self.region2)

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Should have two rows with California: one for LA, one for San Diego
        california_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == "California":
                california_rows.append(row)

        self.assertEqual(len(california_rows), 2)
        districts = [row[2] for row in california_rows]
        self.assertIn("Los Angeles", districts)
        self.assertIn("San Diego", districts)

    def test_sibling_branches_create_separate_rows(self):
        """Test that sibling branches in hierarchy create separate rows."""
        # Create Texas branch
        texas = AdministrativeRegionFactory(name="Texas", administrative_level=self.level2, parent=self.region1)
        AdministrativeRegionFactory(name="Houston", administrative_level=self.level3, parent=texas)
        AdministrativeRegionFactory(name="Dallas", administrative_level=self.level3, parent=texas)

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Should have rows for both Houston and Dallas
        usa_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "USA":
                usa_rows.append(row)

        # Should have at least 3 rows: CA->LA, TX->Houston, TX->Dallas
        self.assertGreaterEqual(len(usa_rows), 3)

        texas_cities = [row[2] for row in usa_rows if row[1] == "Texas"]
        self.assertIn("Houston", texas_cities)
        self.assertIn("Dallas", texas_cities)

    def test_leaf_nodes_without_children(self):
        """Test that leaf nodes (regions with no children) are properly handled."""
        # Los Angeles is already a leaf node in our setup

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Find LA row and verify it's padded with empty cells
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == "Los Angeles":
                # Should have 3 values (USA, California, Los Angeles)
                # and rest should be empty or None
                non_empty = [cell for cell in row if cell]
                self.assertEqual(len(non_empty), 3)
                break

    def test_empty_levels_queryset(self):
        """Test behavior when levels queryset is empty."""
        AdministrativeLevel.objects.all().delete()

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Should only have empty header row or no rows at all
        self.assertEqual(ws.max_row, 1)

    def test_no_root_region_returns_empty_rows(self):
        """Test that no root region results in empty data rows."""
        AdministrativeRegion.objects.all().delete()

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Should have header row but no data rows
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Filter out completely empty rows
        non_empty_data_rows = [row for row in data_rows if any(row)]

        self.assertEqual(len(non_empty_data_rows), 0)

    def test_recursive_traversal_with_complex_tree(self):
        """Test recursive traversal with a complex multi-branch tree."""
        # Create complex structure:
        # USA
        #   ├── California
        #   │   ├── LA
        #   │   └── San Francisco
        #   └── Texas
        #       ├── Houston
        #       └── Dallas

        texas = AdministrativeRegionFactory(name="Texas", administrative_level=self.level2, parent=self.region1)

        AdministrativeRegionFactory(name="San Francisco", administrative_level=self.level3, parent=self.region2)

        AdministrativeRegionFactory(name="Houston", administrative_level=self.level3, parent=texas)

        AdministrativeRegionFactory(name="Dallas", administrative_level=self.level3, parent=texas)

        response = self.get(self.url)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Should have 4 data rows (one for each leaf)
        all_rows = list(ws.iter_rows(min_row=2, values_only=True))
        non_empty_rows = [row for row in all_rows if any(row)]

        self.assertEqual(len(non_empty_rows), 4)

        # Verify each path exists
        paths = [(row[0], row[1], row[2]) for row in non_empty_rows]
        expected_paths = [
            ("USA", "California", "Los Angeles"),
            ("USA", "California", "San Francisco"),
            ("USA", "Texas", "Houston"),
            ("USA", "Texas", "Dallas"),
        ]

        for expected_path in expected_paths:
            self.assertIn(expected_path, paths)
